"""Evaluate pipeline output against a human-expert baseline.

Parses baseline issues from an Excel workbook (Baseline sheet), parses
pipeline output (CSV report or raw manifest + prompts), and produces a
side-by-side comparison with fuzzy-match suggestions and precision/recall
scores.

Usage:
    # Compare a pipeline CSV report against the baseline xlsx
    python entry_points/evaluate.py --baseline Results_02222026.xlsx --report test_results/claude/report_2026-03-11.csv

    # Run pipeline + evaluate in one shot (requires API key)
    python entry_points/evaluate.py --baseline Results_02222026.xlsx --html test_files/home.html --model gpt-4o

    # Compare multiple models against baseline
    python entry_points/evaluate.py --baseline Results_02222026.xlsx --html test_files/home.html --model claude-sonnet-4-20250514 gpt-4o gpt-4.1
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from dataclasses import dataclass, field, asdict
from difflib import SequenceMatcher
from pathlib import Path


# ── Baseline parsing ───────────────────────────────────────────────────────

@dataclass
class BaselineIssue:
    """One issue from the human-expert baseline report."""

    id: int = 0
    element_name: str = ""
    page_title: str = ""
    issue_title: str = ""
    steps_to_reproduce: str = ""
    actual_result: str = ""
    expected_result: str = ""
    recommendation: str = ""
    wcag_sc: str = ""
    category: str = ""
    reported_by: str = ""

    @property
    def text_for_matching(self) -> str:
        """Concatenate key fields into a single string for fuzzy matching."""
        return " ".join([
            self.element_name,
            self.issue_title,
            self.actual_result,
            self.wcag_sc,
        ]).lower()


@dataclass
class PipelineIssue:
    """One issue from a pipeline run (CSV report row)."""

    id: int = 0
    element_name: str = ""
    page_title: str = ""
    issue_title: str = ""
    actual_result: str = ""
    expected_result: str = ""
    recommendation: str = ""
    wcag_sc: str = ""
    category: str = ""
    reported_by: str = ""

    @property
    def text_for_matching(self) -> str:
        """Concatenate key fields into a single string for fuzzy matching."""
        return " ".join([
            self.element_name,
            self.issue_title,
            self.actual_result,
            self.wcag_sc,
        ]).lower()


@dataclass
class MatchResult:
    """A suggested mapping between a pipeline issue and a baseline issue."""

    pipeline_issue: PipelineIssue
    baseline_issue: BaselineIssue | None
    similarity: float = 0.0
    match_type: str = ""  # "strong", "weak", "none"


@dataclass
class EvaluationReport:
    """Full evaluation results for one pipeline run."""

    model: str = ""
    html_file: str = ""
    baseline_count: int = 0
    pipeline_count: int = 0
    matches: list[MatchResult] = field(default_factory=list)
    unmatched_baseline: list[BaselineIssue] = field(default_factory=list)

    @property
    def true_positives(self) -> int:
        return sum(1 for m in self.matches if m.match_type == "strong")

    @property
    def weak_matches(self) -> int:
        return sum(1 for m in self.matches if m.match_type == "weak")

    @property
    def false_positives(self) -> int:
        return sum(1 for m in self.matches if m.match_type == "none")

    @property
    def false_negatives(self) -> int:
        return len(self.unmatched_baseline)

    @property
    def precision(self) -> float:
        tp = self.true_positives
        total = self.pipeline_count
        return tp / total if total > 0 else 0.0

    @property
    def recall(self) -> float:
        tp = self.true_positives
        return tp / self.baseline_count if self.baseline_count > 0 else 0.0

    @property
    def f1(self) -> float:
        p, r = self.precision, self.recall
        return 2 * p * r / (p + r) if (p + r) > 0 else 0.0


# ── Parsing functions ──────────────────────────────────────────────────────

def parse_baseline_xlsx(path: Path, sheet_name: str = "Baseline") -> list[BaselineIssue]:
    """Parse the Baseline sheet from an Excel workbook into BaselineIssue objects."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    # First row is headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    issues = []

    for i, row in enumerate(rows[1:], start=1):
        vals = {headers[j]: (str(row[j]).strip() if row[j] is not None else "")
                for j in range(min(len(headers), len(row)))}

        issue = BaselineIssue(
            id=i,
            element_name=vals.get("element name", ""),
            page_title=vals.get("page title", ""),
            issue_title=vals.get("issue title", ""),
            steps_to_reproduce=vals.get("steps to reproduce", ""),
            actual_result=vals.get("actual result", ""),
            expected_result=vals.get("expected result", ""),
            recommendation=vals.get("recommendation for fix", ""),
            wcag_sc=vals.get("wcag sc", ""),
            category=vals.get("type of change", ""),
            reported_by=vals.get("reported by", ""),
        )
        # Skip empty rows
        if issue.issue_title:
            issues.append(issue)

    wb.close()
    return issues


def parse_pipeline_csv(path: Path) -> list[PipelineIssue]:
    """Parse a pipeline-generated CSV report into PipelineIssue objects."""
    issues = []
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=1):
            issue = PipelineIssue(
                id=i,
                element_name=row.get("element_name", ""),
                page_title=row.get("page_title", ""),
                issue_title=row.get("issue_title", ""),
                actual_result=row.get("actual_result", ""),
                expected_result=row.get("expected_result", ""),
                recommendation=row.get("recommendation", ""),
                wcag_sc=row.get("wcag_sc", ""),
                category=row.get("category", ""),
                reported_by=row.get("reported_by", ""),
            )
            if issue.issue_title:
                issues.append(issue)
    return issues


# ── Fuzzy matching ─────────────────────────────────────────────────────────

STRONG_THRESHOLD = 0.45
WEAK_THRESHOLD = 0.30

# Accessibility concept clusters — issues that use different words to
# describe the same underlying problem.  Each cluster is a set of
# keywords / short phrases; if both issues hit the same cluster the
# similarity score gets a significant boost.
_A11Y_CONCEPT_CLUSTERS: list[set[str]] = [
    {"heading", "h1", "h2", "h3", "h4", "h5", "h6", "heading structure",
     "heading hierarchy", "heading level", "heading markup"},
    {"alt", "alt text", "alternative text", "non-text", "image", "logo",
     "decorative", "informative"},
    {"link", "link text", "link purpose", "cta", "read more", "discernible text"},
    {"landmark", "region", "header", "footer", "nav", "main", "banner",
     "contentinfo"},
    {"label", "form", "input", "field", "placeholder", "required"},
    {"iframe", "embed", "video", "media", "caption", "title"},
    {"focus", "tab order", "focus order", "keyboard"},
    {"carousel", "auto-moving", "pause", "stop", "animation", "slider", "slide"},
    {"announcement", "aria-live", "status message", "dynamic"},
    {"page title", "title", "document title"},
]


def _wcag_overlap(a: str, b: str) -> bool:
    """Check if two WCAG SC strings share any criterion."""
    pattern = r"\d+\.\d+\.\d+"
    set_a = set(re.findall(pattern, a))
    set_b = set(re.findall(pattern, b))
    return bool(set_a & set_b)


def _concept_overlap(text_a: str, text_b: str) -> float:
    """Score how many accessibility concept clusters both texts share.

    Returns a value in [0, 1] based on the fraction of clusters that
    both texts hit.
    """
    a_lower = text_a.lower()
    b_lower = text_b.lower()
    shared = 0
    hit = 0
    for cluster in _A11Y_CONCEPT_CLUSTERS:
        a_hit = any(kw in a_lower for kw in cluster)
        b_hit = any(kw in b_lower for kw in cluster)
        if a_hit or b_hit:
            hit += 1
            if a_hit and b_hit:
                shared += 1
    return shared / hit if hit > 0 else 0.0


def _keyword_overlap(text_a: str, text_b: str) -> float:
    """Jaccard similarity of significant words (length >= 4)."""
    words_a = {w for w in re.findall(r"[a-z]{4,}", text_a.lower())}
    words_b = {w for w in re.findall(r"[a-z]{4,}", text_b.lower())}
    if not words_a or not words_b:
        return 0.0
    intersection = words_a & words_b
    union = words_a | words_b
    return len(intersection) / len(union)


def _similarity(pipeline: PipelineIssue, baseline: BaselineIssue) -> float:
    """Compute a similarity score between a pipeline issue and a baseline issue.

    Combines multiple signals:
    - Keyword overlap (Jaccard on significant words)
    - Accessibility concept cluster overlap
    - WCAG criterion overlap
    - Sequence similarity as tiebreaker
    """
    p_text = pipeline.text_for_matching
    b_text = baseline.text_for_matching

    # Keyword overlap (Jaccard)
    kw_score = _keyword_overlap(p_text, b_text)

    # Concept cluster overlap
    concept_score = _concept_overlap(p_text, b_text)

    # WCAG criterion overlap
    wcag_score = 1.0 if _wcag_overlap(pipeline.wcag_sc, baseline.wcag_sc) else 0.0

    # Sequence similarity (original approach, as tiebreaker)
    seq_score = SequenceMatcher(None, p_text, b_text).ratio()

    # Weighted combination
    return (
        kw_score * 0.25
        + concept_score * 0.35
        + wcag_score * 0.25
        + seq_score * 0.15
    )


def match_issues(
    pipeline_issues: list[PipelineIssue],
    baseline_issues: list[BaselineIssue],
    strong_threshold: float = STRONG_THRESHOLD,
    weak_threshold: float = WEAK_THRESHOLD,
) -> tuple[list[MatchResult], list[BaselineIssue]]:
    """Match pipeline issues to baseline issues using fuzzy matching.

    Uses a greedy best-match approach: for each pipeline issue, find the
    best-matching unmatched baseline issue. Allows multiple pipeline issues
    to map to the same baseline issue (since granularity may differ).

    Returns:
        Tuple of (match_results, unmatched_baseline_issues).
    """
    matches = []
    matched_baseline_ids: set[int] = set()

    # Score all pairs
    for pi in pipeline_issues:
        best_score = 0.0
        best_baseline = None

        for bi in baseline_issues:
            score = _similarity(pi, bi)
            if score > best_score:
                best_score = score
                best_baseline = bi

        if best_score >= strong_threshold and best_baseline is not None:
            match_type = "strong"
            matched_baseline_ids.add(best_baseline.id)
        elif best_score >= weak_threshold and best_baseline is not None:
            match_type = "weak"
            matched_baseline_ids.add(best_baseline.id)
        else:
            match_type = "none"
            best_baseline = None

        matches.append(MatchResult(
            pipeline_issue=pi,
            baseline_issue=best_baseline,
            similarity=best_score,
            match_type=match_type,
        ))

    # Find baseline issues that had no match
    unmatched = [bi for bi in baseline_issues if bi.id not in matched_baseline_ids]

    return matches, unmatched


# ── Evaluation ─────────────────────────────────────────────────────────────

def evaluate(
    pipeline_issues: list[PipelineIssue],
    baseline_issues: list[BaselineIssue],
    model: str = "",
    html_file: str = "",
    strong_threshold: float = STRONG_THRESHOLD,
    weak_threshold: float = WEAK_THRESHOLD,
) -> EvaluationReport:
    """Run full evaluation: match pipeline output against baseline."""
    matches, unmatched = match_issues(
        pipeline_issues, baseline_issues,
        strong_threshold=strong_threshold,
        weak_threshold=weak_threshold,
    )

    return EvaluationReport(
        model=model,
        html_file=html_file,
        baseline_count=len(baseline_issues),
        pipeline_count=len(pipeline_issues),
        matches=matches,
        unmatched_baseline=unmatched,
    )


# ── Output formatting ─────────────────────────────────────────────────────

def format_report(report: EvaluationReport) -> str:
    """Format an EvaluationReport as a human-readable text summary."""
    lines = []
    lines.append("=" * 70)
    lines.append("EVALUATION REPORT")
    lines.append("=" * 70)
    if report.model:
        lines.append(f"Model:          {report.model}")
    if report.html_file:
        lines.append(f"HTML file:      {report.html_file}")
    lines.append(f"Baseline:       {report.baseline_count} issues")
    lines.append(f"Pipeline:       {report.pipeline_count} issues")
    lines.append("")

    # Scores
    lines.append("--- Scores ---")
    lines.append(f"True positives:    {report.true_positives}")
    lines.append(f"Weak matches:      {report.weak_matches}")
    lines.append(f"False positives:   {report.false_positives}")
    lines.append(f"False negatives:   {report.false_negatives}")
    lines.append(f"Precision:         {report.precision:.1%}")
    lines.append(f"Recall:            {report.recall:.1%}")
    lines.append(f"F1 Score:          {report.f1:.1%}")
    lines.append("")

    # Matched issues
    lines.append("--- Matches (pipeline -> baseline) ---")
    for m in sorted(report.matches, key=lambda x: -x.similarity):
        pi = m.pipeline_issue
        tag = {"strong": "Y", "weak": "~", "none": "X"}[m.match_type]
        if m.baseline_issue:
            lines.append(
                f"  [{tag}] ({m.similarity:.0%}) Pipeline #{pi.id}: {pi.issue_title[:60]}"
            )
            lines.append(
                f"       -> Baseline #{m.baseline_issue.id}: {m.baseline_issue.issue_title[:60]}"
            )
        else:
            lines.append(
                f"  [{tag}] Pipeline #{pi.id}: {pi.issue_title[:60]}"
            )
            lines.append(f"       -> (no match)")
    lines.append("")

    # Unmatched baseline issues (false negatives)
    if report.unmatched_baseline:
        lines.append("--- Missed baseline issues (false negatives) ---")
        for bi in report.unmatched_baseline:
            lines.append(f"  Baseline #{bi.id}: {bi.issue_title[:70]}")
            lines.append(f"    WCAG: {bi.wcag_sc}  Element: {bi.element_name}")
        lines.append("")

    return "\n".join(lines)


def save_evaluation_json(report: EvaluationReport, path: Path) -> None:
    """Save evaluation results as JSON for programmatic comparison."""
    data = {
        "model": report.model,
        "html_file": report.html_file,
        "baseline_count": report.baseline_count,
        "pipeline_count": report.pipeline_count,
        "scores": {
            "true_positives": report.true_positives,
            "weak_matches": report.weak_matches,
            "false_positives": report.false_positives,
            "false_negatives": report.false_negatives,
            "precision": round(report.precision, 4),
            "recall": round(report.recall, 4),
            "f1": round(report.f1, 4),
        },
        "matches": [
            {
                "pipeline_id": m.pipeline_issue.id,
                "pipeline_issue": m.pipeline_issue.issue_title,
                "pipeline_wcag": m.pipeline_issue.wcag_sc,
                "baseline_id": m.baseline_issue.id if m.baseline_issue else None,
                "baseline_issue": m.baseline_issue.issue_title if m.baseline_issue else None,
                "baseline_wcag": m.baseline_issue.wcag_sc if m.baseline_issue else None,
                "similarity": round(m.similarity, 4),
                "match_type": m.match_type,
            }
            for m in report.matches
        ],
        "unmatched_baseline": [
            {
                "baseline_id": bi.id,
                "issue_title": bi.issue_title,
                "wcag_sc": bi.wcag_sc,
                "element_name": bi.element_name,
            }
            for bi in report.unmatched_baseline
        ],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print(f"Evaluation JSON saved: {path}")


# ── Multi-model comparison ─────────────────────────────────────────────────

def format_comparison(reports: list[EvaluationReport]) -> str:
    """Format a side-by-side comparison table of multiple evaluation runs."""
    if not reports:
        return "No reports to compare."

    lines = []
    lines.append("=" * 70)
    lines.append("MODEL COMPARISON")
    lines.append("=" * 70)
    lines.append("")

    # Header
    header = f"{'Model':<30} {'TP':>4} {'WM':>4} {'FP':>4} {'FN':>4} {'Prec':>7} {'Rec':>7} {'F1':>7}"
    lines.append(header)
    lines.append("-" * len(header))

    for r in sorted(reports, key=lambda x: -x.f1):
        lines.append(
            f"{r.model:<30} {r.true_positives:>4} {r.weak_matches:>4} "
            f"{r.false_positives:>4} {r.false_negatives:>4} "
            f"{r.precision:>6.1%} {r.recall:>6.1%} {r.f1:>6.1%}"
        )

    lines.append("")
    lines.append("TP=True Positives, WM=Weak Matches, FP=False Positives, FN=False Negatives")
    return "\n".join(lines)


# ── Pipeline runner helper ─────────────────────────────────────────────────

def run_pipeline_and_report(
    html_path: Path,
    model: str,
    output_base: Path,
) -> Path:
    """Run the pipeline + report generation for a given model.

    Returns the path to the generated CSV report.
    """
    output_dir = output_base / model.replace("/", "_")
    report_dir = output_dir / "report"

    # Run pipeline
    cmd = [
        sys.executable, "entry_points/run_pipeline.py",
        "--html", str(html_path),
        "--model", model,
        "--output-dir", str(output_dir),
    ]
    print(f"\n>>> Running pipeline with {model}...")
    result = subprocess.run(cmd, capture_output=True, text=True)
    print(result.stdout)
    if result.returncode != 0:
        print(f"ERROR running pipeline: {result.stderr}")
        return None

    # Generate report
    cmd = [
        sys.executable, "entry_points/generate_report.py",
        "--output-dir", str(output_dir),
        "--report-dir", str(report_dir),
    ]
    print(f">>> Generating report...")
    result = subprocess.run(cmd, capture_output=True, text=True)
    print(result.stdout)
    if result.returncode != 0:
        print(f"ERROR generating report: {result.stderr}")
        return None

    # Find the generated CSV
    csvs = list(report_dir.glob("report_*.csv"))
    if not csvs:
        print(f"ERROR: No report CSV found in {report_dir}")
        return None
    return csvs[0]


# ── CLI ────────────────────────────────────────────────────────────────────

def main():
    """CLI entry point for evaluation."""
    parser = argparse.ArgumentParser(
        description="Evaluate pipeline output against a human-expert baseline."
    )
    parser.add_argument(
        "--baseline", type=Path, required=True,
        help="Path to baseline Excel file (must have a 'Baseline' sheet)",
    )
    parser.add_argument(
        "--baseline-sheet", type=str, default="Baseline",
        help="Sheet name in baseline Excel file (default: Baseline)",
    )
    parser.add_argument(
        "--report", type=Path, default=None,
        help="Path to a pipeline-generated CSV report to evaluate",
    )
    parser.add_argument(
        "--html", type=Path, default=None,
        help="Path to HTML file to run pipeline on (requires API key)",
    )
    parser.add_argument(
        "--model", type=str, nargs="+", default=None,
        help="Model(s) to evaluate (can specify multiple for comparison)",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=Path("./eval_output"),
        help="Base directory for evaluation outputs (default: ./eval_output)",
    )
    parser.add_argument(
        "--strong-threshold", type=float, default=STRONG_THRESHOLD,
        help=f"Similarity threshold for strong match (default: {STRONG_THRESHOLD})",
    )
    parser.add_argument(
        "--weak-threshold", type=float, default=WEAK_THRESHOLD,
        help=f"Similarity threshold for weak match (default: {WEAK_THRESHOLD})",
    )
    args = parser.parse_args()

    # Parse baseline
    print(f"Loading baseline from: {args.baseline}")
    baseline_issues = parse_baseline_xlsx(args.baseline, args.baseline_sheet)
    print(f"  Found {len(baseline_issues)} baseline issues")

    reports: list[EvaluationReport] = []

    if args.report:
        # Evaluate a single existing report
        print(f"Loading pipeline report from: {args.report}")
        pipeline_issues = parse_pipeline_csv(args.report)
        print(f"  Found {len(pipeline_issues)} pipeline issues")

        report = evaluate(
            pipeline_issues, baseline_issues, html_file=str(args.report),
            strong_threshold=args.strong_threshold,
            weak_threshold=args.weak_threshold,
        )
        reports.append(report)

    elif args.html and args.model:
        # Run pipeline for each model, then evaluate
        for model in args.model:
            csv_path = run_pipeline_and_report(args.html, model, args.output_dir)
            if csv_path is None:
                print(f"  Skipping evaluation for {model} (pipeline failed)")
                continue

            pipeline_issues = parse_pipeline_csv(csv_path)
            print(f"  {model}: {len(pipeline_issues)} pipeline issues")

            report = evaluate(
                pipeline_issues, baseline_issues,
                model=model, html_file=str(args.html),
                strong_threshold=args.strong_threshold,
                weak_threshold=args.weak_threshold,
            )
            reports.append(report)
    else:
        parser.error("Provide either --report or both --html and --model")

    # Output results
    args.output_dir.mkdir(parents=True, exist_ok=True)

    for report in reports:
        print("\n" + format_report(report))
        if report.model:
            json_path = args.output_dir / f"eval_{report.model.replace('/', '_')}.json"
        else:
            json_path = args.output_dir / "eval_results.json"
        save_evaluation_json(report, json_path)

    if len(reports) > 1:
        print("\n" + format_comparison(reports))
        # Save comparison
        comp_path = args.output_dir / "comparison.json"
        comp_data = {
            "models": [
                {
                    "model": r.model,
                    "true_positives": r.true_positives,
                    "weak_matches": r.weak_matches,
                    "false_positives": r.false_positives,
                    "false_negatives": r.false_negatives,
                    "precision": round(r.precision, 4),
                    "recall": round(r.recall, 4),
                    "f1": round(r.f1, 4),
                }
                for r in reports
            ],
        }
        with open(comp_path, "w", encoding="utf-8") as f:
            json.dump(comp_data, f, indent=2)
        print(f"Comparison saved: {comp_path}")


if __name__ == "__main__":
    main()
